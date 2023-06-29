import pandas as pd
import re, time
from openpyxl import load_workbook

file_name = r"Listado temas 2023.xlsx"

def submenu():
    print("1.Agregar por consola")
    print("2.Agregar por archivo")
    print("3.Volver al menú principal")

def read_option():
    option = input("Ingresa una opción: ")
    while not option.isdigit() or int(option) < 1 or int(option) > 4:
        option = input("Opción inválida. Ingresa una opción válida: ")
    return option

def validate_input(regex, input_text):
    while not re.match(regex, input_text):
        print("Entrada inválida. Por favor, ingresa un valor válido.")
        input_text = input("Ingrese el valor nuevamente: ")
    return input_text

def add_file():
    print("El archivo ingresado debe respetar la siguiente estructura:\n"
              "Nobre del artista, URL de Spotify, Nombre de la canción,\n"
                "Nombre del álbum, Tipo de álbum, URL, Duración en milisegundos, URL de YouTube, Título, \n"
                "Segundo Nombre del artista, etc..")
    time.sleep(1)
    ruta = input("Escriba la ruta del archivo que desea agregar:")
    try:
        with open(ruta, "r") as file:
            content = file.read()
            content = content.split(",")
            print(content)
            return content
    except Exception as e:
        print("Ocurrió un error:", str(e))
        pass

def write_row(dictionary, sheet):
    row = 2
    for key in dictionary:
        for column in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=column).value == key:
                sheet.cell(row=row, column=column).value = dictionary[key]

def create_empty_row(workbook):
    workbook = workbook
    worksheet = workbook.active
    worksheet.insert_rows(2)

# Opcion 1: Agregar canciones por consola
def add_1(workbook): 

    create_empty_row(workbook)
    sheet = workbook.active

    artist = validate_input(r'^[a-zA-Z0-9\s\_:-]+$', input("Ingrese el nombre del artista: ")) #nombre con cualquier caracter (asd_-:)
    url_spotify = validate_input(r'https:\/\/open\.spotify\.com\/artist\/[A-Za-z0-9]{22}$', input("Ingrese la URL de Spotify: "))  # https://open.spotify.com/artist/1dfeR4HaWDbWqFHLkxsg1d
    track = validate_input(r'^[a-zA-Z0-9\s\_:-]+$', input("Ingrese el nombre de la canción: ")) #nombre con cualquier caracter (asd_-:)
    album = validate_input(r'^[a-zA-Z0-9\s\_:-]+$', input("Ingrese el nombre del álbum: ")) #nombre con cualquier caracter (asd_-:)
    album_type = validate_input(r'^(album|single|compilation)$', input("Ingrese el tipo de álbum: ")) # album, single y compilation
    url = validate_input(r'https:\/\/[A-Za-z0-9\.\/]+$', input("Ingrese la URL: ")) #hay que borrarlo??
    duration_ms = validate_input(r'^\d+$', input("Ingrese la duración en milisegundos: ")) #numero sin puntos
    url_youtube = validate_input(r'https:\/\/www\.youtube\.com\/watch\?v=[A-Za-z0-9\-_]{11}$', input("Ingrese la URL de YouTube: ")) # https://www.youtube.com/watch?v=6Ejga4kJUts
    title = validate_input(r'^[a-zA-Z0-9\s\_:-]+$', input("Ingrese el título: ")) #nombre con cualquier caracter (asd_-:)

    new_row = {
        'Artist': artist,
        'URL_spotify': url_spotify,
        'Track': track,
        'Album': album,
        'Album type': album_type,
        'URL': url, 
        'Duration_ms': int(duration_ms),
        'URL_youtube': url_youtube,
        'title': title
    }

    write_row(new_row, sheet)

    workbook.save(filename = file_name)
    print("El archivo se ha actualizado correctamente.")
    time.sleep(2)   

# Opcion 2: Agregar canciones por archivo
def add_2(workbook, new_content):

    sheet = workbook.active
    new_lines = []
    
    while len(new_content) > 8:

        i = 0
      
        print("Agregando artista: ", new_content[0])
        artist = validate_input(r'^[a-zA-Z0-9\s\_:-]+$', new_content[0])
        print("Agregando URL de Spotify: ", new_content[1])
        url_spotify = validate_input(r'https:\/\/open\.spotify\.com\/artist\/[A-Za-z0-9]{22}$', new_content[1])
        print("Agregando nombre de la canción: ", new_content[2])
        track = validate_input(r'^[a-zA-Z0-9\s\_:-]+$', new_content[2])
        print("Agregando nombre de la cancion: ", new_content[3])
        album = validate_input(r'^[a-zA-Z0-9\s\_:-]+$', new_content[3])
        print("Agregando nombre del álbum: ", new_content[4])
        album_type = validate_input(r'^(album|single|compilation)$', new_content[4])
        print("Agregando URL: ", new_content[5])
        url = validate_input(r'https:\/\/[A-Za-z0-9\.\/]+$', new_content[5])
        print("Agregando duración en milisegundos: ", new_content[6])
        duration_ms = validate_input(r'^\d+$', new_content[6])
        print("Agregando URL de YouTube: ", new_content[7])
        url_youtube = validate_input(r'https:\/\/www\.youtube\.com\/watch\?v=[A-Za-z0-9\-_]{11}$', new_content[7])
        print("Agregando título: ", new_content[8])
        title = validate_input(r'^[a-zA-Z0-9\s\_:-]+$', new_content[8])

        i += 1
        new_content = new_content[9:]
        

        new_line = f"name_line_{i}"

        globals()[new_line] = {
            'Artist': artist,
            'URL_spotify': url_spotify,
            'Track': track,
            'Album': album,
            'Album type': album_type,
            'url': url,
            'Duration_ms': int(duration_ms),
            'URL_youtube': url_youtube,
            'title': title
        }

        new_lines.append(globals()[new_line])

    for line in new_lines:
        create_empty_row(workbook)
        write_row(line, sheet)

    workbook.save(filename = file_name)
    print("El archivo se ha actualizado correctamente.")
    time.sleep(2) 

def run_option4():
    while True:
        submenu()
        option = read_option()
        workbook = load_workbook(filename = file_name)
        if option == "1":
            add_1(workbook)
        elif option == "2":
            new_content = add_file()
            add_2(workbook , new_content)
        elif option == "3":
            break

